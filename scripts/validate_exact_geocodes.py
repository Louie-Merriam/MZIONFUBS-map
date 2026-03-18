#!/usr/bin/env python3

from __future__ import annotations

import argparse
import csv
import importlib.util
import json
import re
import sys
import time
import urllib.parse
import urllib.request
import urllib.error
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT_JSON = ROOT / "mz_data_v5.json"
DEFAULT_CACHE = ROOT / "reports" / "census_validation_cache.json"
DEFAULT_REPORT = ROOT / "reports" / "mz_exact_validation_report.json"
DEFAULT_MISMATCHES = ROOT / "reports" / "mz_exact_validation_mismatches.csv"
DEFAULT_OVERRIDES = ROOT / "reports" / "mz_exact_validation_overrides.json"

DISTANCE_REVIEW_METERS = 100.0
DISTANCE_FIX_METERS = 250.0
ROAD_TYPES = {"ST", "AVE", "RD", "CT", "PL"}
SKIP_PLACE_PATTERNS = [
    re.compile(r"WHITE HOUSE"),
    re.compile(r"CHILDREN'?S HOSPITAL"),
    re.compile(r"TUBERCULOSIS HOSPITAL"),
    re.compile(r"GALLINGER MUNICIPAL HOSPITAL"),
    re.compile(r"SAINT ELIZABETH"),
    re.compile(r"FREEDMEN'?S HOSPITAL"),
]


def load_cleaner() -> Any:
    cleaner_path = ROOT / "scripts" / "build_clean_mz_data.py"
    spec = importlib.util.spec_from_file_location("build_clean_mz_data", cleaner_path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Unable to load cleaner module from {cleaner_path}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


@dataclass
class ValidationTarget:
    query: str
    query_kind: str
    address_key: str
    display_address: str
    state: str
    rows: list[dict[str, Any]]


def load_locations(clean, input_json_path: Path) -> list[Any]:
    rows = clean.current_person_row_map(clean.DEFAULT_WORKBOOK, input_json_path)
    workbook = load_workbook(clean.DEFAULT_WORKBOOK, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
    idx = {header: pos for pos, header in enumerate(headers)}

    locations: list[Any] = []
    for row_number, row, current_person in rows:
        for kind in clean.LOCATION_KEYS:
            loc = clean.build_location(row_number, row, idx, current_person, kind)
            if loc is None:
                continue
            locations.append(loc)

    inferred_locality, locality_fields = clean.infer_locality_from_peers(locations)
    for loc in locations:
        if loc.address and not (loc.city or loc.state or loc.country) and loc.address_key() in inferred_locality:
            loc.inferred_locality = inferred_locality[loc.address_key()]
            if loc.address_key() in locality_fields:
                city, state, country = locality_fields[loc.address_key()]
                loc.city = city
                loc.state = state
                loc.country = country
    return locations


def ambiguous_text(text: str) -> bool:
    normalized = text.upper()
    return (
        "?" in normalized
        or "[?]" in normalized
        or "(?)" in normalized
        or normalized in {"????", "?????", "???"}
    )


def direct_intersection(text: str) -> bool:
    normalized = text.upper()
    if "BETWEEN" in normalized or " BTW " in normalized or " NEAR " in normalized or " VICINITY " in normalized:
        return False
    return " AND " in normalized or " & " in normalized or normalized.startswith("CORNER OF ")


def leading_numbered_fragment(clean, text: str) -> str:
    fragment = clean.leading_address_fragment(text)
    if not fragment:
        return ""
    fragment = re.sub(r"\b(?:BTW|BETWEEN|NEAR|VICINITY OF)\b.*$", "", fragment, flags=re.IGNORECASE)
    fragment = re.sub(
        r"\b([A-Z]|\d+(?:ST|ND|RD|TH)?)\s+AND\s+([A-Z]|\d+(?:ST|ND|RD|TH)?)(?:\s+STS?)?(?:\s+(?:NW|NE|SW|SE))?$",
        "",
        fragment,
        flags=re.IGNORECASE,
    )
    fragment = re.sub(r"\b(?:REAR)\b.*$", "", fragment, flags=re.IGNORECASE)
    fragment = re.sub(r"^\s*REAR\s+", "", fragment, flags=re.IGNORECASE)
    return clean.clean_text(fragment)


def normalized_query(clean, loc: Any) -> tuple[str, str] | None:
    if loc.lat is None or loc.lon is None or not loc.address:
        return None
    if loc.place_name:
        place_key = clean.norm_key(loc.place_name)
        if any(pattern.search(place_key) for pattern in SKIP_PLACE_PATTERNS):
            return None
    if ambiguous_text(loc.address):
        return None
    if clean.has_multiple_address_parts(loc.address):
        return None
    if clean.has_georgetown_historical_name(loc.address):
        return None
    if loc.country and loc.country != "United States":
        return None
    if not (loc.state or loc.city):
        return None

    locality = clean.locality_text(loc.city, loc.state, loc.country or "United States")
    if not locality:
        return None

    if direct_intersection(loc.address):
        query = clean.fix_query_text(clean.dedupe_csv([loc.address, locality]))
        query = re.sub(r",\s*United States$", "", query, flags=re.IGNORECASE)
        return query, "direct_intersection"

    fragment = leading_numbered_fragment(clean, loc.address)
    if fragment:
        query = clean.fix_query_text(clean.dedupe_csv([fragment, locality]))
        query = re.sub(r",\s*United States$", "", query, flags=re.IGNORECASE)
        return query, "numbered_address"

    return None


def load_cache(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    return json.loads(path.read_text())


def save_cache(path: Path, cache: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(cache, indent=2, sort_keys=True))


def geocode_census(query: str, cache: dict[str, Any], sleep_seconds: float) -> dict[str, Any]:
    if query in cache:
        return cache[query]

    params = urllib.parse.urlencode(
        {
            "address": query,
            "benchmark": "Public_AR_Current",
            "format": "json",
        }
    )
    url = f"https://geocoding.geo.census.gov/geocoder/locations/onelineaddress?{params}"
    request = urllib.request.Request(url, headers={"User-Agent": "codex-mz-validator/1.0"})
    try:
        with urllib.request.urlopen(request, timeout=30) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        payload = {
            "error": {
                "code": exc.code,
                "reason": str(exc.reason),
                "query": query,
            }
        }

    cache[query] = payload
    time.sleep(sleep_seconds)
    return payload


def best_match(payload: dict[str, Any], expected_state: str) -> dict[str, Any] | None:
    matches = payload.get("result", {}).get("addressMatches", [])
    if expected_state:
        for match in matches:
            state = str(match.get("addressComponents", {}).get("state") or "")
            if state == expected_state:
                return match
        return None
    for match in matches:
        state = str(match.get("addressComponents", {}).get("state") or "")
        return match
    return matches[0] if matches else None


def build_targets(clean, locations: list[Any]) -> dict[str, ValidationTarget]:
    grouped: dict[str, ValidationTarget] = {}
    for loc in locations:
        candidate = normalized_query(clean, loc)
        if candidate is None:
            continue
        query, query_kind = candidate
        key = json.dumps({"query": query, "state": loc.state, "kind": query_kind}, sort_keys=True)
        entry = {
            "name": loc.person_name,
            "guid": loc.guid,
            "row_number": loc.row_number,
            "kind": loc.kind,
            "display_address": loc.display_address() or "",
            "lat": loc.lat,
            "lon": loc.lon,
            "address_key": loc.address_key(),
        }
        if key not in grouped:
            grouped[key] = ValidationTarget(
                query=query,
                query_kind=query_kind,
                address_key=loc.address_key(),
                display_address=loc.display_address() or "",
                state=loc.state,
                rows=[entry],
            )
        else:
            grouped[key].rows.append(entry)
    return grouped


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "query",
                "query_kind",
                "expected_state",
                "matched_address",
                "matched_state",
                "match_count",
                "distance_meters",
                "name",
                "guid",
                "row_number",
                "kind",
                "display_address",
                "current_lat",
                "current_lon",
                "matched_lat",
                "matched_lon",
                "address_key",
            ],
        )
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def normalize_core(text: str) -> str:
    normalized = text.upper()
    replacements = [
        ("CHAMPLAIN AVE", "CHAMPLAIN ST"),
        ("COCORAN", "CORCORAN"),
        ("CORCRAN", "CORCORAN"),
        ("COREORAN", "CORCORAN"),
        ("FAIRMOUNT", "FAIRMONT"),
        ("KRAEMER", "KRAMER"),
        ("PENNSYLVANNIA", "PENNSYLVANIA"),
        ("POPULAR PL", "POPLAR ST"),
        ("WALLACE", "WALLACH"),
        ("NEW HAMPSHIRE", "NH"),
        ("STREET", "ST"),
        ("AVENUE", "AVE"),
        ("COURT", "CT"),
        ("PLACE", "PL"),
        ("ROAD", "RD"),
    ]
    for old, new in replacements:
        normalized = normalized.replace(old, new)
    normalized = re.sub(r"\bNJ\b", "NEW JERSEY", normalized)
    normalized = re.sub(r",\s*WASHINGTON,\s*DC.*$", "", normalized)
    normalized = re.sub(r",\s*GEORGETOWN,\s*DC.*$", "", normalized)
    normalized = re.sub(r",\s*\d{5}$", "", normalized)
    normalized = normalized.replace("&", " AND ")
    normalized = re.sub(r"[^A-Z0-9/ ]+", " ", normalized)
    normalized = re.sub(r"\b(\d+)(?:ST|ND|RD|TH)\b", r"\1", normalized)
    normalized = re.sub(r"\s+", " ", normalized).strip()
    return normalized


def normalize_intersection_arm(text: str) -> str:
    arm = normalize_core(text)
    arm = re.sub(r"^(?:CORNER OF|CORNER|AT)\s+", "", arm)
    arm = re.sub(r"\bSTRS\b", "ST", arm)
    arm = re.sub(r"\bSTS\b", "ST", arm)
    arm = re.sub(r"\bAV\b", "AVE", arm)
    if re.fullmatch(r"\d+", arm):
        arm = f"{arm} ST"
    elif re.fullmatch(r"[A-Z]", arm):
        arm = f"{arm} ST"
    return re.sub(r"\s+", " ", arm).strip()


def normalized_intersection(text: str) -> tuple[str, str] | None:
    normalized = normalize_core(text)
    normalized = re.sub(r"^(?:CORNER OF|AT)\s+", "", normalized)
    parts = [part.strip() for part in re.split(r"\bAND\b", normalized) if part.strip()]
    if len(parts) != 2:
        return None

    left = normalize_intersection_arm(parts[0])
    right = normalize_intersection_arm(parts[1])
    right_quad = re.search(r"\b(NW|NE|SW|SE)\b", right)
    left_quad = re.search(r"\b(NW|NE|SW|SE)\b", left)
    if right_quad and not left_quad:
        left = f"{left} {right_quad.group(1)}"
    elif left_quad and not right_quad:
        right = f"{right} {left_quad.group(1)}"
    return tuple(sorted([left.strip(), right.strip()]))


def parse_numbered_address(text: str) -> dict[str, Any] | None:
    normalized = normalize_core(text)
    match = re.match(r"^(\d+[A-Z]?)(?:\s+1/2)?\s+(.*)$", normalized)
    if not match:
        return None

    rest = match.group(2)
    quad_match = re.search(r"\b(NW|NE|SW|SE)\b", rest)
    tokens = rest.split()
    core_tokens = [token for token in tokens if token not in ROAD_TYPES and token not in {"NW", "NE", "SW", "SE"}]
    road_type = ""
    for token in reversed(tokens):
        if token in ROAD_TYPES:
            road_type = token
            break
    return {
        "house_number": match.group(1),
        "quad": quad_match.group(1) if quad_match else "",
        "core_tokens": core_tokens,
        "road_type": road_type,
    }


def same_numbered_address_family(query: dict[str, Any], matched: dict[str, Any]) -> bool:
    if query["house_number"] != matched["house_number"]:
        return False
    if query["core_tokens"] != matched["core_tokens"]:
        return False

    if query["road_type"] == matched["road_type"]:
        return True
    if not query["road_type"]:
        return True
    if query["core_tokens"] and query["core_tokens"][0].isdigit():
        return {query["road_type"], matched["road_type"]} <= {"", "ST"}
    return False


def trusted_census_match(row: dict[str, Any]) -> bool:
    if row["query_kind"] == "direct_intersection":
        query_parts = normalized_intersection(row["address_key"])
        match_parts = normalized_intersection(row["matched_address"])
        if query_parts is None:
            return False
        if query_parts != match_parts:
            # Georgetown references without an explicit quadrant can still be trusted
            # when the matched result resolves both arms to NW.
            context = f"{row.get('query', '')} {row.get('display_address', '')}".upper()
            if "GEORGETOWN" not in context or match_parts is None:
                return False
            query_core = [tuple(token for token in arm.split() if token not in {"NW", "NE", "SW", "SE"}) for arm in query_parts]
            match_core = [tuple(token for token in arm.split() if token not in {"NW", "NE", "SW", "SE"}) for arm in match_parts]
            if sorted(query_core) != sorted(match_core):
                return False
            if not all(arm.endswith(" NW") for arm in match_parts):
                return False
        if re.search(r"\b(NW|NE|SW|SE)\b", row["address_key"]):
            return True
        return int(row.get("match_count") or 0) == 1

    query_match = parse_numbered_address(row["address_key"])
    matched_match = parse_numbered_address(row["matched_address"])
    if not query_match or not matched_match:
        return False
    if not same_numbered_address_family(query_match, matched_match):
        return False
    if query_match["quad"] and matched_match["quad"]:
        return query_match["quad"] == matched_match["quad"]
    if query_match["quad"] or matched_match["quad"]:
        return int(row.get("match_count") or 0) == 1
    return int(row.get("match_count") or 0) == 1


def main() -> None:
    parser = argparse.ArgumentParser(description="Validate exact modern MZ/FUBS geocodes against the U.S. Census geocoder.")
    parser.add_argument("--input-json", type=Path, default=DEFAULT_INPUT_JSON)
    parser.add_argument("--cache", type=Path, default=DEFAULT_CACHE)
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT)
    parser.add_argument("--mismatches", type=Path, default=DEFAULT_MISMATCHES)
    parser.add_argument("--overrides", type=Path, default=DEFAULT_OVERRIDES)
    parser.add_argument("--sleep-seconds", type=float, default=0.05)
    args = parser.parse_args()

    clean = load_cleaner()
    locations = load_locations(clean, args.input_json)
    targets = build_targets(clean, locations)
    cache = load_cache(args.cache)

    stats = Counter()
    mismatches: list[dict[str, Any]] = []
    suggested_overrides: dict[str, dict[str, float]] = {}
    trusted_overrides: dict[str, dict[str, float]] = {}
    progress = 0

    for target in targets.values():
        payload = geocode_census(target.query, cache, args.sleep_seconds)
        progress += 1
        if progress % 25 == 0:
            save_cache(args.cache, cache)
        if progress % 50 == 0:
            print(json.dumps({"validated_queries": progress, "total_queries": len(targets)}), flush=True)

        if payload.get("error"):
            stats["query_errors"] += len(target.rows)
            continue

        match = best_match(payload, target.state)
        if match is None:
            stats[f"{target.query_kind}_no_match"] += len(target.rows)
            continue

        match_count = len(payload.get("result", {}).get("addressMatches", []))
        matched_address = match.get("matchedAddress") or ""
        matched_state = str(match.get("addressComponents", {}).get("state") or "")
        matched_lon = float(match["coordinates"]["x"])
        matched_lat = float(match["coordinates"]["y"])

        for row in target.rows:
            current = (float(row["lat"]), float(row["lon"]))
            matched = (matched_lat, matched_lon)
            distance = clean.haversine_meters(current, matched)
            stats["validated_locations"] += 1
            stats[f"{target.query_kind}_validated"] += 1
            row_payload = {
                "query": target.query,
                "query_kind": target.query_kind,
                "expected_state": target.state,
                "matched_address": matched_address,
                "matched_state": matched_state,
                "match_count": match_count,
                "distance_meters": round(distance, 1),
                "name": row["name"],
                "guid": row["guid"],
                "row_number": row["row_number"],
                "kind": row["kind"],
                "display_address": row["display_address"],
                "current_lat": row["lat"],
                "current_lon": row["lon"],
                "matched_lat": matched_lat,
                "matched_lon": matched_lon,
                "address_key": row["address_key"],
            }
            if distance > DISTANCE_REVIEW_METERS:
                stats["review_distance_locations"] += 1
                if row["address_key"] and trusted_census_match(row_payload):
                    trusted_overrides[row["address_key"]] = {
                        "lat": matched_lat,
                        "lon": matched_lon,
                    }
            if distance > DISTANCE_FIX_METERS:
                stats["fix_distance_locations"] += 1
                mismatches.append(row_payload)
                if row["address_key"]:
                    suggested_overrides[row["address_key"]] = {
                        "lat": matched_lat,
                        "lon": matched_lon,
                    }

    save_cache(args.cache, cache)
    args.report.parent.mkdir(parents=True, exist_ok=True)
    report = {
        "input_json": str(args.input_json),
        "exact_query_count": len(targets),
        "validated_location_count": stats["validated_locations"],
        "review_distance_locations": stats["review_distance_locations"],
        "fix_distance_locations": stats["fix_distance_locations"],
        "suggested_override_count": len(suggested_overrides),
        "trusted_override_count": len(trusted_overrides),
        "stats": dict(stats),
        "review_distance_threshold_meters": DISTANCE_REVIEW_METERS,
        "fix_distance_threshold_meters": DISTANCE_FIX_METERS,
    }
    args.report.write_text(json.dumps(report, indent=2, sort_keys=True))
    write_csv(args.mismatches, mismatches)
    args.overrides.write_text(json.dumps(trusted_overrides, indent=2, sort_keys=True))
    print(json.dumps(report, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
