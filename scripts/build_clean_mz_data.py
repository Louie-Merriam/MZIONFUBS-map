#!/usr/bin/env python3

import argparse
import csv
import json
import math
import re
import time
import urllib.parse
import urllib.request
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = Path("/Users/lmerriam27/Downloads/MZFUBS_tbl_DECEDENT_2026-Jan-12_1822/MZFUBS_tbl_DECEDENT_2026-Jan-12_1822.xlsx")
DEFAULT_INPUT_JSON = ROOT / "mz_data_v4.json"
DEFAULT_OUTPUT_JSON = ROOT / "mz_data_v5.json"
DEFAULT_STATE_POLYGONS = ROOT / "state_polygons_simplified.json"
DEFAULT_CACHE = ROOT / "reports" / "geocode_cache.json"
DEFAULT_REPORT = ROOT / "reports" / "mz_geocode_cleaning_report.json"
DEFAULT_MANUAL_REVIEW = ROOT / "reports" / "mz_manual_review.csv"
DEFAULT_EXACT_VALIDATION_OVERRIDES = ROOT / "reports" / "mz_exact_validation_overrides.json"

LOCATION_KEYS = ["birth", "residence", "enslavement", "worship", "death"]

FIELD_SPECS = {
    "birth": {
        "place_name": None,
        "address_cols": ["BirthPlaceAddress"],
        "city_col": "BirthPlaceCity",
        "state_col": "BirthPlaceState",
        "country_col": "BirthPlaceCountry",
        "lat_col": "BirthPlaceLatitude",
        "lon_col": "BirthPlaceLongitude",
    },
    "residence": {
        "place_name": None,
        "address_cols": ["ResidenceAddress"],
        "city_col": "ResidenceCity",
        "state_col": "ResidenceState",
        "country_col": "ResidenceCountry",
        "lat_col": "ResidenceLatitude",
        "lon_col": "ResidenceLongitude",
    },
    "enslavement": {
        "place_name": "EnslavementPlaceName",
        "address_cols": ["EnslavementAddress"],
        "city_col": "EnslavementCity",
        "state_col": "EnslavementState",
        "country_col": "EnslavementCountry",
        "lat_col": "EnslavementLatitude",
        "lon_col": "EnslavementLongitude",
    },
    "worship": {
        "place_name": "PlaceOfWorshipName",
        "address_cols": ["POWAddressModern", "POWAddressHistorical"],
        "city_col": "POWCity",
        "state_col": "POWState",
        "country_col": "POWCountry",
        "lat_col": "POWLatitude",
        "lon_col": "POWLongitude",
    },
    "death": {
        "place_name": "DeathPlaceName",
        "address_cols": ["DeathPlaceAddress"],
        "city_col": "DeathPlaceCity",
        "state_col": "DeathPlaceState",
        "country_col": "DeathPlaceCountry",
        "lat_col": "DeathPlaceLatitude",
        "lon_col": "DeathPlaceLongitude",
    },
}

US_STATE_NAME_TO_CODE = {
    "ALABAMA": "AL",
    "ALASKA": "AK",
    "ARIZONA": "AZ",
    "ARKANSAS": "AR",
    "CALIFORNIA": "CA",
    "COLORADO": "CO",
    "CONNECTICUT": "CT",
    "DELAWARE": "DE",
    "DISTRICT OF COLUMBIA": "DC",
    "FLORIDA": "FL",
    "GEORGIA": "GA",
    "HAWAII": "HI",
    "IDAHO": "ID",
    "ILLINOIS": "IL",
    "INDIANA": "IN",
    "IOWA": "IA",
    "KANSAS": "KS",
    "KENTUCKY": "KY",
    "LOUISIANA": "LA",
    "MAINE": "ME",
    "MARYLAND": "MD",
    "MASSACHUSETTS": "MA",
    "MICHIGAN": "MI",
    "MINNESOTA": "MN",
    "MISSISSIPPI": "MS",
    "MISSOURI": "MO",
    "MONTANA": "MT",
    "NEBRASKA": "NE",
    "NEVADA": "NV",
    "NEW HAMPSHIRE": "NH",
    "NEW JERSEY": "NJ",
    "NEW MEXICO": "NM",
    "NEW YORK": "NY",
    "NORTH CAROLINA": "NC",
    "NORTH DAKOTA": "ND",
    "OHIO": "OH",
    "OKLAHOMA": "OK",
    "OREGON": "OR",
    "PENNSYLVANIA": "PA",
    "RHODE ISLAND": "RI",
    "SOUTH CAROLINA": "SC",
    "SOUTH DAKOTA": "SD",
    "TENNESSEE": "TN",
    "TEXAS": "TX",
    "UTAH": "UT",
    "VERMONT": "VT",
    "VIRGINIA": "VA",
    "WASHINGTON": "WA",
    "WEST VIRGINIA": "WV",
    "WISCONSIN": "WI",
    "WYOMING": "WY",
}
US_STATE_CODE_TO_NAME = {v: k for k, v in US_STATE_NAME_TO_CODE.items()}
COUNTRY_TOKENS = {"UNITED STATES", "UNITED STATES OF AMERICA", "USA", "US"}
DC_CENTER = (38.8950368, -77.0365427)
GEORGETOWN_CENTER = (38.9051994, -77.062785)
DC_VIEWBOX = (-77.12, 38.995, -76.91, 38.79)
GEORGETOWN_VIEWBOX = (-77.075, 38.921, -77.041, 38.899)

STATE_CENTROIDS = {
    "AL": (32.806671, -86.79113),
    "AK": (61.370716, -152.404419),
    "AZ": (33.729759, -111.431221),
    "AR": (34.969704, -92.373123),
    "CA": (36.116203, -119.681564),
    "CO": (39.059811, -105.311104),
    "CT": (41.597782, -72.755371),
    "DE": (39.318523, -75.507141),
    "DC": (38.9072, -77.0369),
    "FL": (27.766279, -81.686783),
    "GA": (33.040619, -83.643074),
    "HI": (21.094318, -157.498337),
    "ID": (44.240459, -114.478828),
    "IL": (40.349457, -88.986137),
    "IN": (39.849426, -86.258278),
    "IA": (42.011539, -93.210526),
    "KS": (38.5266, -96.726486),
    "KY": (37.66814, -84.670067),
    "LA": (31.169546, -91.867805),
    "ME": (44.693947, -69.381927),
    "MD": (39.045755, -76.641271),
    "MA": (42.230171, -71.530106),
    "MI": (43.326618, -84.536095),
    "MN": (45.694454, -93.900192),
    "MS": (32.741646, -89.678696),
    "MO": (38.456085, -92.288368),
    "MT": (46.921925, -110.454353),
    "NE": (41.12537, -98.268082),
    "NV": (38.313515, -117.055374),
    "NH": (43.452492, -71.563896),
    "NJ": (40.298904, -74.521011),
    "NM": (34.840515, -106.248482),
    "NY": (42.165726, -74.948051),
    "NC": (35.630066, -79.806419),
    "ND": (47.528912, -99.784012),
    "OH": (40.388783, -82.764915),
    "OK": (35.565342, -96.928917),
    "OR": (44.572021, -122.070938),
    "PA": (40.590752, -77.209755),
    "RI": (41.680893, -71.51178),
    "SC": (33.856892, -80.945007),
    "SD": (44.299782, -99.438828),
    "TN": (35.747845, -86.692345),
    "TX": (31.054487, -97.563461),
    "UT": (40.150032, -111.862434),
    "VT": (44.045876, -72.710686),
    "VA": (37.769337, -78.169968),
    "WA": (47.400902, -121.490494),
    "WV": (38.491226, -80.954453),
    "WI": (44.268543, -89.616508),
    "WY": (42.755966, -107.30249),
}

SOURCE_BAD_CENTROIDS = {
    "VA": (37.1232245, -78.4927721),
    "MD": (39.5162401, -76.9382069),
    "PA": (40.9699889, -77.7278831),
    "NC": (35.6729639, -79.0392919),
    "SC": (33.6874388, -80.4363743),
    "WV": (38.4758406, -80.8408415),
}

QUERY_FIXES = {
    "BALSTON": "BALLSTON",
    "BARNSVILLE": "BARNESVILLE",
    "DENINDDY": "DINWIDDIE",
    "DRAINESVILLE": "DRANESVILLE",
    "FAIRFAX COURT HOUSE": "FAIRFAX COURTHOUSE",
    "GETTYSBURG, PA, PA": "GETTYSBURG, PA, UNITED STATES",
    "LAPLATA": "LA PLATA",
    "LEESBURGN": "LEESBURG",
    "MANASAS": "MANASSAS",
    "PHILADEPHIA": "PHILADELPHIA",
    "PHILADELPHA": "PHILADELPHIA",
    "RICHMOND CITY": "RICHMOND",
    "SAINT PAULS": "ST. PAULS",
    "SHEPERDSTOWN": "SHEPHERDSTOWN",
    "SNOW COURT": "SNOWS COURT",
    "STANTON": "STAUNTON",
    "VANPELT": "VAN PELT",
    "WEST MORELAND": "WESTMORELAND",
    "WILKES BARRE": "WILKES-BARRE",
}

VERIFIED_DISPLAY_FIXES = {
    "ALEXANDRIA, VA, UNITED STATES": {
        "lat": 38.8051095,
        "lon": -77.0470229,
    },
    "NORBECK, MD, UNITED STATES": {
        "lat": 39.1094423,
        "lon": -77.0769636,
    },
    "NORFOLK, VA, UNITED STATES": {
        "lat": 36.8493695,
        "lon": -76.2899539,
    },
    "TUBERCULOSIS HOSPITAL, SHEPHERDSTOWN, SHEPERDSTOWN, WY, UNITED STATES": {
        "city": "Shepherdstown",
        "state": "WV",
        "country": "United States",
        "lat": 39.4300996,
        "lon": -77.8041610,
    },
}

OUTSIDE_STATE_FALSE_POSITIVES = {
    "ALEXANDRIA, VA, UNITED STATES",
}

GEORGETOWN_HISTORICAL_RENAMES = [
    (r"\bNORTH ST(?:REET)?(?:\s+NW)?\b", "26th Street NW"),
    (r"\bMONROE ST(?:REET)?(?:\s+NW)?\b", "27th Street NW"),
    (r"\bMONTGOMERY ST(?:REET)?(?:\s+NW)?\b", "28th Street NW"),
    (r"\bGREENE ST(?:REET)?(?:\s+NW)?\b", "29th Street NW"),
    (r"\bWASHINGTON ST(?:REET)?(?:\s+NW)?\b", "30th Street NW"),
    (r"\bCONGRESS ST(?:REET)?(?:\s+NW)?\b", "31st Street NW"),
    (r"\bVALLEY ST(?:REET)?(?:\s+NW)?\b", "32nd Street NW"),
    (r"\bHIGH ST(?:REET)?(?:\s+NW)?\b", "Wisconsin Avenue NW"),
    (r"\bMARKET ST(?:REET)?(?:\s+NW)?\b", "33rd Street NW"),
    (r"\bFREDERICK ST(?:REET)?(?:\s+NW)?\b", "34th Street NW"),
    (r"\bFAYETTE ST(?:REET)?(?:\s+NW)?\b", "35th Street NW"),
    (r"\bLINGAN ST(?:REET)?(?:\s+NW)?\b", "36th Street NW"),
    (r"\bWARREN ST(?:REET)?(?:\s+NW)?\b", "37th Street NW"),
    (r"\bWATER ST(?:REET)?(?:\s+NW)?\b", "K Street NW"),
    (r"\bBRIDGE ST(?:REET)?(?:\s+NW)?\b", "M Street NW"),
    (r"\bFALLS ST(?:REET)?(?:\s+NW)?\b", "M Street NW"),
    (r"\bOLIVE ST(?:REET)?(?:\s+NW)?\b", "Prospect Street NW"),
    (r"\bGAY(?:\s*\([^)]*\))?\s+STREET(?:\s+NW)?\b", "N Street NW"),
    (r"\bGAY(?:\s*\([^)]*\))?\s+ST(?:REET)?(?:\s+NW)?\b", "N Street NW"),
    (r"\bFIRST ST(?:REET)?(?:\s+NW)?\b", "N Street NW"),
    (r"\bBEALL(?:\s+ST(?:REET)?)?(?:\s*\([^)]*\))?(?:\s+NW)?\b", "O Street NW"),
    (r"\bSECOND ST(?:REET)?(?:\s+NW)?\b", "O Street NW"),
    (r"\bWEST ST(?:REET)?(?:\s+NW)?\b", "P Street NW"),
    (r"\bTHIRD ST(?:REET)?(?:\s+NW)?\b", "P Street NW"),
    (r"\bWEST P STREET\b", "P Street NW"),
    (r"\b4TH ST(?:REET)?(?:\s+NW)?(?:\s*\([^)]*\))?\b", "Volta Place NW"),
    (r"\bSTODD(?:A|E)RT ST(?:REET)?(?:\s+NW)?\b", "Q Street NW"),
    (r"\bFIFTH ST(?:REET)?(?:\s+NW)?\b", "Q Street NW"),
    (r"\bMILL ST(?:REET)?(?:\s+NW)?\b", "27th Street NW"),
    (r"\bROCK ST(?:REET)?(?:\s+NW)?\b", "27th Street NW"),
    (r"\bSIXTH ST(?:REET)?(?:\s+NW)?\b", "Dent Place NW"),
    (r"\bSEVENTH ST(?:REET)?(?:\s+NW)?\b", "Reservoir Road NW"),
    (r"\bROAD ST(?:REET)?(?:\s+NW)?\b", "R Street NW"),
    (r"\b8TH ST(?:REET)?(?:\s+NW)?\b", "R Street NW"),
    (r"\bJEFFERSON (?:ST|STREET|AVE|AVENUE)(?:\s+NW)?\b", "Thomas Jefferson Street NW"),
]

GEORGETOWN_OUTLIER_OVERRIDES = [
    (r"\bBEALL(?:\s+ST(?:REET)?)?\b", (38.9084951, -77.0593482)),
    (r"\bWEST ST(?:REET)?\b", (38.9086667, -77.0715563)),
    (r"\bNORTH ST(?:REET)?\b", (38.9108837, -77.0547912)),
    (r"\bMONROE ST(?:REET)?\b", (38.9095985, -77.0554969)),
    (r"\bMONTGOMERY ST(?:REET)?\b", (38.9087494, -77.0572017)),
    (r"\bGREENE ST(?:REET)?\b", (38.9083288, -77.0581586)),
    (r"\bWASHINGTON ST(?:REET)?\b", (38.905219213882, -77.059167697721)),
    (r"\bCONGRESS ST(?:REET)?\b", (38.905192213963, -77.061183697095)),
    (r"\bVALLEY ST(?:REET)?\b", (38.911758, -77.0641446)),
    (r"\bMARKET ST(?:REET)?\b", (38.9096445, -77.0661012)),
    (r"\bHIGH ST(?:REET)?\b", (38.9158362, -77.0678491)),
    (r"\bBRIDGE ST(?:REET)?\b", (38.9052340, -77.0586415)),
    (r"\bFAYETTE ST(?:REET)?\b", (38.9140819, -77.0693050)),
    (r"\bFREDERICK ST(?:REET)?\b", (38.9104643, -77.0678609)),
    (r"\bGAY(?:\s*\([^)]*\))?\s+ST(?:REET)?\b", (38.9068107, -77.0624425)),
    (r"\b8TH ST(?:REET)?\b", (38.9135422, -77.0638911)),
    (r"\b4TH ST(?:REET)?\b", (38.9097718, -77.0668751)),
    (r"\bJEFFERSON (?:ST|STREET|AVE|AVENUE)\b", (38.9035719, -77.0602296)),
    (r"\bOLIVE ST(?:REET)?\b", (38.9059426, -77.0575317)),
]

# Verified overrides are intentionally narrow: only rows with a defensible
# modern equivalent or a clearly wrong upstream locality are patched here.
VERIFIED_ADDRESS_FIXES = {
    "1038 JEFFERSON AVENUE": {
        "lat": 38.9035719,
        "lon": -77.0602296,
    },
    "1038 JEFFERSON AVENUE / STREET": {
        "lat": 38.9035719,
        "lon": -77.0602296,
    },
    "1141 21 ST L AND M STS": {
        "lat": 38.90476743467192,
        "lon": -77.04631602,
    },
    "1220 WOOD STREET": {
        "city": "Philadelphia",
        "state": "PA",
        "country": "United States",
        "lat": 39.9583380,
        "lon": -75.1589033,
    },
    "12460 MAYHURST LANE": {
        "lat": 38.2328309,
        "lon": -78.1174545,
    },
    "1253 22ND ST NW": {
        "lat": 38.906762459644,
        "lon": -77.048730572348,
    },
    "1002 18TH ST": {
        "lat": 38.9026688,
        "lon": -77.0417897,
    },
    "1064 30TH ST": {
        "lat": 38.9044221,
        "lon": -77.0593105,
    },
    "12ST AND M STREET NW": {
        "lat": 38.905642212637,
        "lon": -77.028081707374,
    },
    "1418 28TH (MONTGOMERY) STREET, NEAR O (BEALL) STREET": {
        "lat": 38.9087494,
        "lon": -77.0572017,
    },
    "1418 MONTGOMERY STREET NEAR BEALL STREET": {
        "lat": 38.9087494,
        "lon": -77.0572017,
    },
    "1418 MONTGOMERY STREET, NEAR BEALL STREET": {
        "lat": 38.9087494,
        "lon": -77.0572017,
    },
    "1414 27TH ST": {
        "lat": 38.9089277,
        "lon": -77.0559980,
    },
    "1417 WEST 28TH STREET": {
        "lat": 38.9090582,
        "lon": -77.0569383,
    },
    "1515 27TH ST": {
        "lat": 38.9097018,
        "lon": -77.0553133,
    },
    "1523 26TH ST NW": {
        "city": "Washington",
        "state": "DC",
        "country": "United States",
        "lat": 38.9099569,
        "lon": -77.0546221,
    },
    "1615 33RD ST NW": {
        "lat": 38.910937123693,
        "lon": -77.066039845123,
    },
    "1626 VALLEY STREET": {
        "lat": 38.9111636,
        "lon": -77.0640726,
    },
    "1670 VALLEY STREET NW": {
        "lat": 38.9123525,
        "lon": -77.0642166,
    },
    "19TH ST BTW L AND M STREETS": {
        "lat": 38.9047017134,
        "lon": -77.043453702645,
    },
    "231 MASSACHUSETTS AVE NW": {
        "lat": 38.8992228,
        "lon": -77.0138840,
    },
    "14 STODDARD ST": {
        "city": "Georgetown",
        "state": "DC",
        "country": "United States",
        "lat": 38.9105206,
        "lon": -77.0599388,
    },
    "21ST ST BTW L AND M STREETS": {
        "lat": 38.9047167,
        "lon": -77.0466399,
    },
    "21 STODDARD ST": {
        "city": "Georgetown",
        "state": "DC",
        "country": "United States",
        "lat": 38.9105206,
        "lon": -77.0599388,
    },
    "23RD ST BETWEEN H AND I STREETS": {
        "lat": 38.9001387145695,
        "lon": -77.0501294508455,
    },
    "23RD STREET NW BETWEEN L AND M STREETS": {
        "lat": 38.9044992136895,
        "lon": -77.050111450584,
    },
    "2117 N DENINDDY STREET": {
        "lat": 38.8954480,
        "lon": -77.1283933,
    },
    "2018 E ST NW": {
        "lat": 38.8960988,
        "lon": -77.0453161,
    },
    "2040 E ST NW": {
        "lat": 38.8961061,
        "lon": -77.0456399,
    },
    "2104 35TH ST NW": {
        "lat": 38.918021,
        "lon": -77.069326,
    },
    "2108 E ST NW": {
        "lat": 38.8961019,
        "lon": -77.0468999,
    },
    "2122 E ST NW": {
        "lat": 38.8961017,
        "lon": -77.0470478,
    },
    "24 4TH STREET (VOLTA STREET / PLACE)": {
        "lat": 38.9097718,
        "lon": -77.0668751,
    },
    "24 4TH STREET (VOLTA STREET)": {
        "lat": 38.9097718,
        "lon": -77.0668751,
    },
    "24 4TH STREET (VOLTA STREET/PLACE)": {
        "lat": 38.9097718,
        "lon": -77.0668751,
    },
    "2411 SNOW COURT": {
        "lat": 38.9015677,
        "lon": -77.0528308,
    },
    "2718 POPLAR PLACE": {
        "lat": 38.90890225,
        "lon": -77.05654947,
    },
    "2724 P STREET NW": {
        "city": "Georgetown",
        "state": "DC",
        "country": "United States",
        "lat": 38.9091952,
        "lon": -77.0564557,
    },
    "27TH AND N STREETS": {
        "lat": 38.906875213423,
        "lon": -77.055801698667,
    },
    "2812 OLIVE AVE NW": {
        "city": "Washington",
        "state": "DC",
        "country": "United States",
        "lat": 38.9059426,
        "lon": -77.0575317,
    },
    "2910 O (BEALL) STREET NW": {
        "lat": 38.9084139,
        "lon": -77.0583925,
    },
    "816 G ST SW": {
        "lat": 38.8813580,
        "lon": -77.0234348,
    },
    "41 4TH STREET NW": {
        "lat": 38.9097718,
        "lon": -77.0668751,
    },
    "5 STODDARD ST": {
        "city": "Georgetown",
        "state": "DC",
        "country": "United States",
        "lat": 38.9105206,
        "lon": -77.0599388,
    },
    "514 RICKETTS COURT": {
        "lat": 38.8987855412,
        "lon": -77.050777885,
    },
    "455 WEST P STREET": {
        "lat": 38.9093454,
        "lon": -77.0583594,
    },
    "6TH AND H STREETS, NW": {
        "lat": 38.8997458,
        "lon": -77.0201773,
    },
    "8 STODDARD ST": {
        "city": "Georgetown",
        "state": "DC",
        "country": "United States",
        "lat": 38.9105206,
        "lon": -77.0599388,
    },
    "STODDARD REAR OF MONROE": {
        "city": "Georgetown",
        "state": "DC",
        "country": "United States",
        "lat": 38.9105208,
        "lon": -77.0551813,
    },
    "STODDARD ST": {
        "city": "Georgetown",
        "state": "DC",
        "country": "United States",
        "lat": 38.9105206,
        "lon": -77.0599388,
    },
    "STODDARD ST BTW MILL AND ROCK STS": {
        "city": "Georgetown",
        "state": "DC",
        "country": "United States",
        "lat": 38.9105208,
        "lon": -77.0551813,
    },
    "STODDARD ST NEAR MILL ST": {
        "city": "Georgetown",
        "state": "DC",
        "country": "United States",
        "lat": 38.9105208,
        "lon": -77.0551813,
    },
    "STODDARD ST NEAR NORTH ST": {
        "city": "Georgetown",
        "state": "DC",
        "country": "United States",
        "lat": 38.910574212646,
        "lon": -77.054964698702,
    },
    "STODDARD ST NEAR ROCK CREEK": {
        "city": "Georgetown",
        "state": "DC",
        "country": "United States",
        "lat": 38.9105208,
        "lon": -77.0551813,
    },
    "EAST STODDARD ST BTW MILL AND MONROE STS": {
        "city": "Georgetown",
        "state": "DC",
        "country": "United States",
        "lat": 38.9105208,
        "lon": -77.0551813,
    },
    "VALLEY ST BTW ROAD AND STODDARD STS": {
        "city": "Georgetown",
        "state": "DC",
        "country": "United States",
        "lat": 38.9120502,
        "lon": -77.0640322,
    },
    "VALLEY ST BTW STODDARD AND 8TH": {
        "city": "Georgetown",
        "state": "DC",
        "country": "United States",
        "lat": 38.9120502,
        "lon": -77.0640322,
    },
    "ALLEY BTW 24TH AND 25TH, M AND N STREETS NW, DC": {
        "lat": 38.906215963427,
        "lon": -77.052351324782,
    },
    "ALLEY, ALLEY BTW 24TH AND 25TH, M AND N STREETS NW, DC": {
        "lat": 38.906215963427,
        "lon": -77.052351324782,
    },
    "CONGRESS STREET ALLEY": {
        "lat": 38.9054508002,
        "lon": -77.0610942,
    },
    "HUGHES ALLEY BTW 25TH AND 26TH J AND K STS": {
        "lat": 38.902801,
        "lon": -77.054012,
    },
    "NEAR 20TH AND M STREETS": {
        "lat": 38.905660213261,
        "lon": -77.044898202137,
    },
    "RESIDENCE, 2125 K STREET NW": {
        "lat": 38.9026561,
        "lon": -77.0472345,
    },
    "VICINITY OF 30TH AND M STREETS (8003 M STREET, AS OF 1939)": {
        "lat": 38.905219213882,
        "lon": -77.059167697721,
    },
    "14 WEST MASSACHUSETTS AVE": {
        "lat": 38.898070822659,
        "lon": -77.010696907905,
    },
    "1255 WARD PLACE, NW": {
        "lat": 38.9063466,
        "lon": -77.0478530,
    },
    "1314 28TH ST": {
        "lat": 38.9073553,
        "lon": -77.0571750,
    },
    "1324 27TH": {
        "lat": 38.9074874,
        "lon": -77.0559623,
    },
    "1326 27TH ST": {
        "lat": 38.9075734,
        "lon": -77.0559367,
    },
    "1100 ALABAMA AND MASSACHUSETTS AVE SE": {
        "lat": 38.844385231633,
        "lon": -76.990537696818,
    },
    "1348 28TH ST BTW O AND DUMBARTON ST": {
        "lat": 38.907982257271,
        "lon": -77.057145561288,
    },
    "1728 17TH STREET R AND S": {
        "lat": 38.913446204915,
        "lon": -77.038554839343,
    },
    "21ST AND P ST": {
        "lat": 38.909638212524,
        "lon": -77.046635201353,
    },
    "CONNECTICUT AVE AND M ST": {
        "lat": 38.90566021312,
        "lon": -77.041127203311,
    },
    "CHERRY ALLEY AND CISSEL ALLEY (NOW CECILY PLACE, NEAR THE CANAL)": {
        "lat": 38.9032953,
        "lon": -77.0659997,
    },
}

HISTORICAL_FACILITY_OVERRIDES = [
    (
        re.compile(r"CHILDREN'?S HOSPITAL.*13TH AND V STREET NW"),
        {
            "lat": 38.918647210073,
            "lon": -77.029620706132,
        },
    ),
    (
        re.compile(r"TUBERCULOSIS HOSPITAL.*(?:13TH|14TH) AND UPSHUR"),
        {
            "lat": 38.941867205454,
            "lon": -77.031215954182,
        },
    ),
    (
        re.compile(r"REAR LAWN OF THE WHITE HOUSE.*1600 PENNSYLVANIA AVENUE"),
        {
            "lat": 38.8976997,
            "lon": -77.03655315,
        },
    ),
    (
        re.compile(r"FREEDMEN'?S HOSPITAL.*13TH AND R STREETS NW"),
        {
            "lat": 38.91260821129,
            "lon": -77.029620706469,
        },
    ),
    (
        re.compile(r"FREEDMEN'?S HOSPITAL.*BRYANT AND 6TH STREETS NW"),
        {
            "lat": 38.920096209449,
            "lon": -77.020724208782,
        },
    ),
    (
        re.compile(r"GALLINGER MUNICIPAL HOSPITAL"),
        {
            "lat": 38.884555214996,
            "lon": -76.977258724488,
        },
    ),
    (
        re.compile(r"^19TH AND C ST(?:REET)?S? SE, WASHINGTON, DC, UNITED STATES$"),
        {
            "lat": 38.885347214837,
            "lon": -76.977263224438,
        },
    ),
    (
        re.compile(r"SAINT ELIZABETH'?S(?: HOSPITAL)?.*(?:2700 AND 2701 MARTIN LUTHER KING JR|1100 ALABAMASSACHUSETTS AVENUE SE|NICHOLS AVENUE)"),
        {
            "lat": 38.848418338288,
            "lon": -76.996067557514,
        },
    ),
]

if DEFAULT_EXACT_VALIDATION_OVERRIDES.exists():
    EXTERNAL_VERIFIED_ADDRESS_FIXES = json.loads(DEFAULT_EXACT_VALIDATION_OVERRIDES.read_text())
else:
    EXTERNAL_VERIFIED_ADDRESS_FIXES = {}


@dataclass
class Location:
    row_number: int
    guid: str
    person_name: str
    kind: str
    place_name: str
    address: str
    city: str
    state: str
    country: str
    current_lat: float | None
    current_lon: float | None
    raw_lat: float | None
    raw_lon: float | None
    current_json_address: str | None
    lat: float | None = None
    lon: float | None = None
    inferred_locality: str = ""
    issues: list[str] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)
    changed: bool = False
    attempted_queries: list[str] = field(default_factory=list)
    geocode_source: str = ""

    def display_address(self) -> str | None:
        parts: list[str] = []
        if self.place_name:
            parts.append(self.place_name)
        if self.address:
            parts.append(self.address)
        locality = locality_text(self.city, self.state, self.country)
        if locality:
            parts.append(locality)
        if not parts:
            return None
        return dedupe_csv(parts)

    def query_address(self) -> str | None:
        parts: list[str] = []
        if self.address:
            parts.append(self.address)
        locality = locality_text(self.city, self.state, self.country) or self.inferred_locality
        if locality:
            parts.append(locality)
        if self.place_name and not self.address:
            parts.insert(0, self.place_name)
        if not parts:
            return None
        return dedupe_csv(parts)

    def address_key(self) -> str:
        return norm_key(self.address)

    def structured_key(self) -> str:
        return " | ".join(
            part for part in [norm_key(self.address), norm_key(self.city), norm_key(self.state), norm_key(self.country)] if part
        )

    def locality_key(self) -> str:
        return " | ".join(part for part in [norm_key(self.city), norm_key(self.state), norm_key(self.country)] if part)


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return text.strip(" ,")


def normalize_name(value: Any) -> str:
    return clean_text(value)


def norm_key(value: str) -> str:
    return re.sub(r"\s+", " ", clean_text(value).upper().replace(".", "")).strip(" ,")


def clean_country_value(value: str) -> tuple[str, str | None]:
    text = clean_text(value)
    if not text:
        return "", None
    normalized = norm_key(text)
    compact = re.sub(r"[^A-Z]", "", normalized)
    if compact in {"UNITEDSTATES", "UNITEDSTATESOFAMERICA", "US", "USA"}:
        return "United States", None
    if compact.endswith("UNITEDSTATES"):
        prefix = compact[: -len("UNITEDSTATES")]
        if prefix in US_STATE_CODE_TO_NAME:
            return "United States", prefix
    if normalized in COUNTRY_TOKENS:
        return "United States", None
    return text, None


def state_code_for(value: str) -> str | None:
    token = norm_key(value).replace("STATE OF ", "").strip()
    if not token:
        return None
    if token in US_STATE_CODE_TO_NAME:
        return token
    return US_STATE_NAME_TO_CODE.get(token)


def fix_query_text(text: str) -> str:
    fixed = clean_text(text)
    fixed = re.sub(r"^\s*NO\.?\s*", "", fixed, flags=re.IGNORECASE)
    for label in ["Residence", "Alley"]:
        fixed = re.sub(rf"^(?:{label}\s*,\s*)+", "", fixed, flags=re.IGNORECASE)
    for old, new in QUERY_FIXES.items():
        fixed = re.sub(rf"\b{re.escape(old)}\b", new, fixed, flags=re.IGNORECASE)
    fixed = re.sub(r"\bALABAMASSACHUSETTS\b", "ALABAMA", fixed, flags=re.IGNORECASE)
    fixed = re.sub(r"\bUNITED STATES(?:[A-Z]+|UNITED STATES)\b", "United States", fixed, flags=re.IGNORECASE)
    return clean_text(fixed)


def locality_text(city: str, state: str, country: str) -> str:
    parts = [part for part in [clean_text(city), clean_text(state), clean_text(country)] if part]
    if not parts:
        return ""
    return dedupe_csv(parts)


def dedupe_csv(parts: list[str]) -> str:
    out: list[str] = []
    for part in parts:
        text = clean_text(part)
        if not text:
            continue
        if not out or norm_key(out[-1]) != norm_key(text):
            out.append(text)
    return ", ".join(out)


def parse_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def clean_date_text(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return clean_text(value)


def parse_year_value(value: Any) -> int | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return int(value)
    text = clean_text(value)
    if text.isdigit():
        return int(text)
    match = re.search(r"\b(\d{4})\b", text)
    if match:
        return int(match.group(1))
    return None


def parse_age_years(value: Any) -> int | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return int(value)
    text = clean_text(value)
    match = re.search(r"(\d+)", text)
    if match:
        return int(match.group(1))
    return None


def build_timeline(row: tuple[Any, ...], idx: dict[str, int]) -> dict[str, Any] | None:
    birth_year = parse_year_value(row[idx["BirthYear"]])
    death_year = parse_year_value(row[idx["DeathYear"]])
    birth_date = clean_date_text(row[idx["BirthDate"]])
    death_date = clean_date_text(row[idx["DeathDate"]])
    age_at_death_years = parse_age_years(row[idx["AgeDeathYrs"]])

    start_year = birth_year
    start_source = "birth_year" if birth_year is not None else None
    if start_year is None and death_year is not None and age_at_death_years is not None and 0 <= age_at_death_years <= 120:
        estimated = death_year - age_at_death_years
        if estimated <= death_year:
            start_year = estimated
            start_source = "estimated_from_age_at_death"

    if start_year is not None and death_year is not None and start_year > death_year:
        start_year = None
        start_source = None

    end_year = death_year
    end_source = "death_year" if death_year is not None else None
    if end_year is None and start_year is not None:
        end_year = start_year + 50
        end_source = "approximate_50_year_cap"

    if all(value in (None, "") for value in [birth_year, death_year, birth_date, death_date, start_year]):
        return None

    label = ""
    if start_year is not None and death_year is not None:
        prefix = "c. " if start_source == "estimated_from_age_at_death" and birth_year is None else ""
        label = f"{prefix}{start_year}-{death_year}"
    elif start_year is not None and end_source == "approximate_50_year_cap":
        prefix = "c. " if start_source == "estimated_from_age_at_death" and birth_year is None else ""
        label = f"{prefix}{start_year}-c. {end_year}"
    elif birth_year is not None:
        label = f"b. {birth_year}"
    elif death_year is not None:
        label = f"d. {death_year}"

    payload = {
        "birthYear": birth_year,
        "birthDate": birth_date or None,
        "deathYear": death_year,
        "deathDate": death_date or None,
        "startYear": start_year,
        "endYear": end_year,
        "startYearSource": start_source,
        "endYearSource": end_source,
        "estimatedBirthYear": bool(start_source == "estimated_from_age_at_death"),
        "estimatedDeathYear": bool(end_source == "approximate_50_year_cap"),
        "ageAtDeathYears": age_at_death_years,
        "filterable": bool(start_year is not None),
        "label": label or None,
    }
    return payload


def coords_equal(a: tuple[float, float] | None, b: tuple[float, float] | None, tolerance: float = 1e-6) -> bool:
    if not a or not b:
        return False
    return abs(a[0] - b[0]) <= tolerance and abs(a[1] - b[1]) <= tolerance


def haversine_meters(a: tuple[float, float], b: tuple[float, float]) -> float:
    lat1, lon1 = a
    lat2, lon2 = b
    r = 6371000.0
    p1 = math.radians(lat1)
    p2 = math.radians(lat2)
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    aa = math.sin(dlat / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dlon / 2) ** 2
    return 2 * r * math.asin(math.sqrt(aa))


def point_in_ring(lat: float, lon: float, ring: list[list[float]]) -> bool:
    inside = False
    count = len(ring)
    if count < 3:
        return False
    for i in range(count):
        y1, x1 = ring[i]
        y2, x2 = ring[(i + 1) % count]
        if (x1 > lon) != (x2 > lon):
            y_intersection = (y2 - y1) * (lon - x1) / (x2 - x1) + y1
            if lat < y_intersection:
                inside = not inside
    return inside


def point_in_state(lat: float, lon: float, state_code: str, polygons: dict[str, list[list[list[float]]]]) -> bool:
    for ring in polygons.get(state_code, []):
        if point_in_ring(lat, lon, ring):
            return True
    return False


def row_has_any_coord_pair(row: tuple[Any, ...], idx: dict[str, int]) -> bool:
    for kind in LOCATION_KEYS:
        spec = FIELD_SPECS[kind]
        if row[idx[spec["lat_col"]]] not in (None, "") and row[idx[spec["lon_col"]]] not in (None, ""):
            return True
    return False


def looks_like_street(text: str) -> bool:
    normalized = norm_key(text)
    if not normalized:
        return False
    if re.search(r"\b\d+(?:ST|ND|RD|TH)\b", normalized):
        return True
    return any(
        token in normalized
        for token in [
            " STREET",
            " ST ",
            " AVE",
            " AVENUE",
            " ROAD",
            " RD ",
            " LANE",
            " LN ",
            " COURT",
            " CT ",
            " ALLEY",
            " PLACE",
            " PL ",
            " SQUARE",
            " SQ ",
            " BETWEEN ",
            " BTW ",
            " ALLEY",
        ]
    )


def looks_like_dc_local_address(text: str) -> bool:
    normalized = norm_key(text)
    if not normalized:
        return False
    return bool(
        re.search(r"\b(NW|NE|SW|SE)\b", normalized)
        or "GEORGETOWN" in normalized
        or "SNOWS CT" in normalized
        or "DUMBARTON" in normalized
        or "MONROE ST" in normalized
        or "P STREET" in normalized
        or "Q STREET" in normalized
    )


def has_multiple_address_parts(text: str) -> bool:
    normalized = norm_key(text)
    return any(token in normalized for token in [";", " THEN ", " AND AT ", " LIVED ON ", " BRIDGE AT "])


def leading_address_fragment(text: str) -> str:
    candidate = re.sub(r"^\s*NO\.?\s*", "", clean_text(text), flags=re.IGNORECASE)
    match = re.match(r"^\d+[A-Z]?(?:[-/]\d+)?\s+[^,.;]+", candidate, flags=re.IGNORECASE)
    if not match:
        return ""
    fragment = clean_text(match.group(0))
    return fragment if looks_like_street(fragment) else ""


def georgetown_historical_variants(loc: Location) -> list[str]:
    if norm_key(loc.city) != "GEORGETOWN" or not loc.address or has_multiple_address_parts(loc.address):
        return []

    seeds = {clean_text(loc.address)}
    fragment = leading_address_fragment(loc.address)
    if fragment:
        seeds.add(fragment)

    expanded_seeds = set(seeds)
    for seed in list(seeds):
        stripped = clean_text(re.sub(r"\([^)]*\)", "", seed))
        if stripped:
            expanded_seeds.add(stripped)

    variants: list[str] = []
    for seed in expanded_seeds:
        current = seed
        changed = False
        for pattern, replacement in GEORGETOWN_HISTORICAL_RENAMES:
            replaced = re.sub(pattern, replacement, current, flags=re.IGNORECASE)
            if replaced != current:
                current = clean_text(replaced)
                changed = True
        if changed and norm_key(current) not in {norm_key(seed), norm_key(loc.address)}:
            variants.append(current)

    deduped: list[str] = []
    seen: set[str] = set()
    for variant in variants:
        key = norm_key(variant)
        if key and key not in seen:
            seen.add(key)
            deduped.append(variant)
    return deduped


def has_georgetown_historical_name(text: str) -> bool:
    if not text:
        return False
    return any(re.search(pattern, text, flags=re.IGNORECASE) for pattern, _ in GEORGETOWN_HISTORICAL_RENAMES)


def centroid_for_city(city: str, state: str) -> tuple[float, float] | None:
    if state != "DC":
        return None
    city_key = norm_key(city)
    if city_key == "GEORGETOWN":
        return GEORGETOWN_CENTER
    if city_key == "WASHINGTON":
        return DC_CENTER
    return None


def suspicious_small_pair(lat: float | None, lon: float | None) -> bool:
    if lat is None or lon is None:
        return False
    return abs(lat) <= 12 and abs(lon) <= 12 and float(int(lat)) == lat and float(int(lon)) == lon and lat >= 0 and lon >= 0


def dc_distance_km(lat: float, lon: float) -> float:
    return haversine_meters((lat, lon), DC_CENTER) / 1000.0


def repair_components(city: str, state: str, country: str) -> tuple[str, str, str, list[str]]:
    notes: list[str] = []
    city = clean_text(city)
    state = clean_text(state)
    country = clean_text(country)

    country, moved_state = clean_country_value(country)
    if moved_state and not state:
        state = moved_state
        notes.append("moved_state_from_country")

    if norm_key(city) in {"DC", "DISTRICT OF COLUMBIA"} and norm_key(state) == "WASHINGTON":
        city = "Washington"
        state = "DC"
        country = "United States"
        notes.append("swapped_dc_washington_fields")

    if norm_key(country) == "WASHINGTON" and norm_key(state) in {"DC", "DISTRICT OF COLUMBIA"}:
        state = "DC"
        country = "United States"
        notes.append("fixed_country_washington_suffix")

    state_code = state_code_for(state)
    if state_code:
        if state != state_code:
            notes.append("normalized_state_code")
        state = state_code
        if not country:
            country = "United States"
            notes.append("filled_us_country_from_state")

    country, moved_state = clean_country_value(country)
    if moved_state and not state:
        state = moved_state
        notes.append("moved_state_from_country")

    if state and not country and state_code_for(state):
        country = "United States"
        notes.append("filled_us_country_from_state")

    if country and country != "United States" and state == "DC":
        country = "United States"
        notes.append("normalized_us_country_variant")

    return city, state, country, notes


def build_location(
    row_number: int,
    row: tuple[Any, ...],
    idx: dict[str, int],
    current_person: dict[str, Any],
    kind: str,
) -> Location | None:
    spec = FIELD_SPECS[kind]
    current = current_person.get(kind)
    if current is None:
        return None

    place_name = clean_text(row[idx[spec["place_name"]]]) if spec["place_name"] else ""
    address = ""
    for col in spec["address_cols"]:
        candidate = clean_text(row[idx[col]])
        if candidate:
            address = candidate
            break

    city = clean_text(row[idx[spec["city_col"]]])
    state = clean_text(row[idx[spec["state_col"]]])
    country = clean_text(row[idx[spec["country_col"]]])
    city, state, country, locality_notes = repair_components(city, state, country)

    current_address = current.get("address") if isinstance(current, dict) else None
    if not address and not city and not state and not country and current_address:
        address = clean_text(current_address)
        locality_notes.append("used_current_json_address_fallback")

    loc = Location(
        row_number=row_number,
        guid=clean_text(row[idx["Decedent_GUID"]]),
        person_name=normalize_name(row[idx["FullName"]]),
        kind=kind,
        place_name=place_name,
        address=address,
        city=city,
        state=state,
        country=country,
        current_lat=parse_float(current.get("lat")),
        current_lon=parse_float(current.get("lon")),
        raw_lat=parse_float(row[idx[spec["lat_col"]]]),
        raw_lon=parse_float(row[idx[spec["lon_col"]]]),
        current_json_address=clean_text(current_address),
    )
    loc.lat = loc.current_lat
    loc.lon = loc.current_lon
    loc.notes.extend(locality_notes)
    return loc


def classify_issue(loc: Location, polygons: dict[str, Any], conflict_keys: set[str]) -> list[str]:
    issues: list[str] = []
    display_key = norm_key(loc.display_address() or "")
    if loc.lat is None or loc.lon is None:
        issues.append("missing_coords")
        return issues
    if not (-90 <= loc.lat <= 90 and -180 <= loc.lon <= 180):
        issues.append("out_of_range")
    if suspicious_small_pair(loc.lat, loc.lon) or (
        suspicious_small_pair(loc.raw_lat, loc.raw_lon) and (loc.lat is None or loc.lon is None or suspicious_small_pair(loc.lat, loc.lon))
    ):
        issues.append("placeholder_coords")
    if loc.state and not loc.city and not loc.address and loc.country == "United States":
        if loc.state in STATE_CENTROIDS and not coords_equal((loc.lat, loc.lon), STATE_CENTROIDS[loc.state]):
            issues.append("state_only_not_at_centroid")
    if loc.city and loc.state in STATE_CENTROIDS and not loc.address and coords_equal((loc.lat, loc.lon), STATE_CENTROIDS[loc.state]):
        issues.append("city_at_state_centroid")
    if loc.city and loc.state in SOURCE_BAD_CENTROIDS and not loc.address and coords_equal((loc.lat, loc.lon), SOURCE_BAD_CENTROIDS[loc.state]):
        issues.append("city_at_state_centroid")
    city_centroid = centroid_for_city(loc.city, loc.state)
    if loc.address and looks_like_street(loc.address) and city_centroid and coords_equal((loc.lat, loc.lon), city_centroid):
        issues.append("street_at_city_centroid")
    if (
        norm_key(loc.city) == "GEORGETOWN"
        and loc.address
        and has_georgetown_historical_name(loc.address)
        and haversine_meters((loc.lat, loc.lon), GEORGETOWN_CENTER) > 2000
    ):
        issues.append("georgetown_historical_outlier")
    if loc.address and not loc.city and not loc.state and looks_like_dc_local_address(loc.address) and dc_distance_km(loc.lat, loc.lon) > 100:
        issues.append("street_only_far_from_dc")
    if (
        loc.state
        and loc.country == "United States"
        and loc.state in polygons
        and display_key not in OUTSIDE_STATE_FALSE_POSITIVES
        and not point_in_state(loc.lat, loc.lon, loc.state, polygons)
    ):
        issues.append("outside_expected_state")
    if loc.structured_key() and loc.structured_key() in conflict_keys:
        issues.append("conflicting_address_coords")
    return issues


def geocode_candidates(loc: Location) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []
    base_query = loc.query_address()
    locality = locality_text(loc.city, loc.state, loc.country) or loc.inferred_locality
    expected_state = loc.state if loc.state in US_STATE_CODE_TO_NAME else None
    expect_us = loc.country == "United States" or bool(expected_state)
    georgetown_context = norm_key(loc.city) == "GEORGETOWN"

    def add_candidate(text: str, with_place_name: bool = False, use_georgetown_bbox: bool = False) -> None:
        query = fix_query_text(text)
        if not query:
            return
        candidate = {
            "query": query,
            "expected_state": expected_state,
            "expect_us": expect_us,
        }
        if use_georgetown_bbox:
            candidate["viewbox"] = GEORGETOWN_VIEWBOX
        elif expected_state == "DC" or "WASHINGTON, DC" in query.upper() or "GEORGETOWN, DC" in query.upper():
            candidate["viewbox"] = DC_VIEWBOX
        if candidate not in candidates:
            candidates.append(candidate)

    if loc.address and looks_like_dc_local_address(loc.address) and not locality:
        add_candidate(dedupe_csv([loc.address, "Washington, DC, United States"]))
    fragment = leading_address_fragment(loc.address) if loc.address else ""
    if fragment and locality and norm_key(fragment) != norm_key(loc.address):
        add_candidate(dedupe_csv([fragment, locality]), use_georgetown_bbox=georgetown_context)
    if georgetown_context and base_query:
        add_candidate(base_query, use_georgetown_bbox=True)
    for variant in georgetown_historical_variants(loc):
        add_candidate(dedupe_csv([variant, locality or "Georgetown, DC, United States"]), use_georgetown_bbox=True)
    if loc.place_name and base_query:
        add_candidate(dedupe_csv([loc.place_name, base_query]), with_place_name=True, use_georgetown_bbox=georgetown_context)
    if base_query and not georgetown_context and not (loc.address and looks_like_dc_local_address(loc.address) and not locality):
        add_candidate(base_query)
    if locality and not loc.address:
        add_candidate(locality)
    if loc.address and loc.inferred_locality and loc.inferred_locality != locality:
        add_candidate(dedupe_csv([loc.address, loc.inferred_locality]), use_georgetown_bbox=georgetown_context)
    return candidates


def apply_verified_location_fix(loc: Location) -> bool:
    display_key = norm_key(loc.display_address() or "")
    for pattern, fix in HISTORICAL_FACILITY_OVERRIDES:
        if pattern.search(display_key):
            target_coords = (fix["lat"], fix["lon"])
            if loc.lat is None or loc.lon is None or not coords_equal((loc.lat, loc.lon), target_coords):
                loc.lat, loc.lon = target_coords
                loc.changed = True
                loc.notes.append("applied_historical_facility_override")
            return True

    fix = VERIFIED_DISPLAY_FIXES.get(norm_key(loc.display_address() or ""))
    if loc.address:
        external_fix = EXTERNAL_VERIFIED_ADDRESS_FIXES.get(loc.address_key())
        if external_fix is not None:
            fix = external_fix
        elif fix is None:
            fix = VERIFIED_ADDRESS_FIXES.get(loc.address_key())
    if fix is None:
        return False

    updated = False
    for field_name in ["address", "city", "state", "country"]:
        target = fix.get(field_name)
        if target is not None and getattr(loc, field_name) != target:
            setattr(loc, field_name, target)
            updated = True

    target_coords = (fix["lat"], fix["lon"])
    if loc.lat is None or loc.lon is None or not coords_equal((loc.lat, loc.lon), target_coords):
        loc.lat, loc.lon = target_coords
        updated = True

    if updated:
        loc.changed = True
        loc.notes.append("applied_verified_override")
    return updated


def apply_georgetown_outlier_override(loc: Location) -> bool:
    if norm_key(loc.city) != "GEORGETOWN" or not loc.address or loc.lat is None or loc.lon is None:
        return False
    if haversine_meters((loc.lat, loc.lon), GEORGETOWN_CENTER) <= 2000:
        return False

    for pattern, target in GEORGETOWN_OUTLIER_OVERRIDES:
        if not re.search(pattern, loc.address, flags=re.IGNORECASE):
            continue
        if not coords_equal((loc.lat, loc.lon), target):
            loc.lat, loc.lon = target
            loc.changed = True
            loc.notes.append("applied_georgetown_outlier_override")
        return True
    return False


def load_cache(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    return json.loads(path.read_text())


def save_cache(path: Path, cache: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(cache, indent=2, sort_keys=True))


def geocode_query(candidate: dict[str, Any], cache: dict[str, Any], sleep_seconds: float) -> list[dict[str, Any]]:
    key = json.dumps(candidate, sort_keys=True)
    if key in cache:
        return cache[key]

    params = {
        "q": candidate["query"],
        "format": "jsonv2",
        "limit": 5,
        "addressdetails": 1,
    }
    if candidate.get("viewbox"):
        params["viewbox"] = ",".join(str(value) for value in candidate["viewbox"])
        params["bounded"] = "1"

    url = "https://nominatim.openstreetmap.org/search?" + urllib.parse.urlencode(params)
    request = urllib.request.Request(url, headers={"User-Agent": "codex-geocode-cleaner/1.0"})
    with urllib.request.urlopen(request, timeout=30) as response:
        results = json.loads(response.read().decode("utf-8"))

    cache[key] = results
    time.sleep(sleep_seconds)
    return results


def acceptable_result(result: dict[str, Any], candidate: dict[str, Any], polygons: dict[str, Any]) -> bool:
    try:
        lat = float(result["lat"])
        lon = float(result["lon"])
    except (KeyError, TypeError, ValueError):
        return False

    expected_state = candidate.get("expected_state")
    if expected_state:
        if expected_state in polygons and not point_in_state(lat, lon, expected_state, polygons):
            return False

    if candidate.get("expect_us"):
        address = result.get("address") or {}
        country_code = str(address.get("country_code") or "").lower()
        if country_code and country_code != "us":
            return False

    return True


def apply_state_centroid_fix(loc: Location) -> bool:
    if loc.state and not loc.city and not loc.address and loc.country == "United States" and loc.state in STATE_CENTROIDS:
        target = STATE_CENTROIDS[loc.state]
        if not coords_equal((loc.lat, loc.lon) if loc.lat is not None and loc.lon is not None else None, target):
            loc.lat, loc.lon = target
            loc.changed = True
            loc.notes.append("replaced_with_state_centroid")
        return True
    return False


def infer_locality_from_peers(locations: list[Location]) -> tuple[dict[str, str], dict[str, tuple[str, str, str]]]:
    inferred_queries: dict[str, str] = {}
    locality_by_address: dict[str, Counter[str]] = defaultdict(Counter)
    fields_by_address: dict[str, tuple[str, str, str]] = {}
    for loc in locations:
        if not loc.address or not (loc.city or loc.state or loc.country):
            continue
        locality = locality_text(loc.city, loc.state, loc.country)
        if not locality:
            continue
        key = loc.address_key()
        locality_by_address[key][locality] += 1
        fields_by_address.setdefault(key, (loc.city, loc.state, loc.country))

    for key, counter in locality_by_address.items():
        if len(counter) == 1:
            inferred_queries[key] = next(iter(counter))
    return inferred_queries, fields_by_address


def consensus_coords(locations: list[Location]) -> dict[str, tuple[float, float]]:
    buckets: dict[str, Counter[tuple[float, float]]] = defaultdict(Counter)
    for loc in locations:
        if not loc.structured_key() or loc.lat is None or loc.lon is None:
            continue
        if "placeholder_coords" in loc.issues:
            continue
        buckets[loc.structured_key()][(round(loc.lat, 6), round(loc.lon, 6))] += 1

    out: dict[str, tuple[float, float]] = {}
    for key, counter in buckets.items():
        if len(counter) == 1:
            out[key] = next(iter(counter))
    return out


def conflicting_keys(locations: list[Location]) -> set[str]:
    buckets: dict[str, list[tuple[float, float]]] = defaultdict(list)
    for loc in locations:
        if not loc.address or not loc.structured_key() or loc.lat is None or loc.lon is None:
            continue
        buckets[loc.structured_key()].append((loc.lat, loc.lon))

    out: set[str] = set()
    for key, coords in buckets.items():
        distinct = {(round(lat, 6), round(lon, 6)) for lat, lon in coords}
        if len(distinct) < 2:
            continue
        pairs = list(distinct)
        max_distance = 0.0
        for i in range(len(pairs)):
            for j in range(i + 1, len(pairs)):
                max_distance = max(max_distance, haversine_meters(pairs[i], pairs[j]))
        if max_distance > 50:
            out.add(key)
    return out


def current_person_row_map(workbook_path: Path, input_json_path: Path) -> list[tuple[int, tuple[Any, ...], dict[str, Any]]]:
    current_json = json.loads(input_json_path.read_text())["people"]
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
    idx = {header: pos for pos, header in enumerate(headers)}
    exported_rows: list[tuple[int, tuple[Any, ...]]] = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row_has_any_coord_pair(row, idx):
            exported_rows.append((row_number, row))
    if len(exported_rows) != len(current_json):
        raise RuntimeError(f"Workbook export rows ({len(exported_rows)}) do not match JSON rows ({len(current_json)})")
    return [(row_number, row, person) for (row_number, row), person in zip(exported_rows, current_json)]


def build_clean_dataset(
    workbook_path: Path,
    input_json_path: Path,
    state_polygons_path: Path,
    geocode_live: bool,
    cache_path: Path,
    sleep_seconds: float,
) -> tuple[dict[str, Any], dict[str, Any], list[dict[str, Any]]]:
    polygons = json.loads(state_polygons_path.read_text())
    rows = current_person_row_map(workbook_path, input_json_path)

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
    idx = {header: pos for pos, header in enumerate(headers)}

    people: list[dict[str, Any]] = []
    all_locations: list[Location] = []

    for row_number, row, current_person in rows:
        person = {"name": normalize_name(row[idx["FullName"]])}
        person_locations: list[Location] = []
        for kind in LOCATION_KEYS:
            loc = build_location(row_number, row, idx, current_person, kind)
            if loc is None:
                continue
            person_locations.append(loc)
            all_locations.append(loc)
        people.append(person)

    inferred_locality, locality_fields = infer_locality_from_peers(all_locations)
    for loc in all_locations:
        if loc.address and not (loc.city or loc.state or loc.country) and loc.address_key() in inferred_locality:
            loc.inferred_locality = inferred_locality[loc.address_key()]
            if loc.address_key() in locality_fields:
                city, state, country = locality_fields[loc.address_key()]
                loc.city = city
                loc.state = state
                loc.country = country
            loc.notes.append("inferred_locality_from_peer_address")

    for loc in all_locations:
        apply_state_centroid_fix(loc)

    for loc in all_locations:
        apply_verified_location_fix(loc)
        apply_georgetown_outlier_override(loc)

    conflicts = conflicting_keys(all_locations)
    for loc in all_locations:
        loc.issues = classify_issue(loc, polygons, conflicts)

    consensus = consensus_coords(all_locations)
    for loc in all_locations:
        key = loc.structured_key()
        if key in consensus and ("placeholder_coords" in loc.issues or "street_at_city_centroid" in loc.issues):
            target = consensus[key]
            if not coords_equal((loc.lat, loc.lon), target):
                loc.lat, loc.lon = target
                loc.changed = True
                loc.notes.append("replaced_with_consensus_coords")

    conflicts = conflicting_keys(all_locations)
    for loc in all_locations:
        loc.issues = classify_issue(loc, polygons, conflicts)

    cache = load_cache(cache_path)
    query_stats = Counter()
    if geocode_live:
        for loc in all_locations:
            needs_geocode = any(
                issue in loc.issues
                for issue in [
                    "placeholder_coords",
                    "city_at_state_centroid",
                    "georgetown_historical_outlier",
                    "street_at_city_centroid",
                    "street_only_far_from_dc",
                    "outside_expected_state",
                    "conflicting_address_coords",
                ]
            )
            if not needs_geocode:
                continue

            for candidate in geocode_candidates(loc):
                loc.attempted_queries.append(candidate["query"])
                results = geocode_query(candidate, cache, sleep_seconds)
                query_stats["queries"] += 1
                accepted = next((result for result in results if acceptable_result(result, candidate, polygons)), None)
                if accepted is None:
                    continue
                new_lat = float(accepted["lat"])
                new_lon = float(accepted["lon"])
                if loc.lat is None or loc.lon is None or haversine_meters((loc.lat, loc.lon), (new_lat, new_lon)) > 5:
                    loc.lat = new_lat
                    loc.lon = new_lon
                    loc.changed = True
                    loc.geocode_source = candidate["query"]
                    loc.notes.append("updated_from_live_geocode")
                query_stats["accepted"] += 1
                break

    save_cache(cache_path, cache)

    conflicts = conflicting_keys(all_locations)
    for loc in all_locations:
        loc.issues = classify_issue(loc, polygons, conflicts)

    report = {
        "source_workbook": str(workbook_path),
        "input_json": str(input_json_path),
        "output_json": str(DEFAULT_OUTPUT_JSON),
        "people_count": len(rows),
        "location_counts": {},
        "changed_locations": 0,
        "dropped_locations": 0,
        "issue_counts": Counter(),
        "notes_counts": Counter(),
        "query_stats": dict(query_stats),
        "timeline_counts": Counter(),
    }
    manual_review: list[dict[str, Any]] = []

    output_people: list[dict[str, Any]] = []
    row_iter = iter(all_locations)
    all_locations_by_person: dict[tuple[int, str], list[Location]] = defaultdict(list)
    for loc in all_locations:
        all_locations_by_person[(loc.row_number, loc.person_name)].append(loc)

    for row_number, row, current_person in rows:
        key = (row_number, normalize_name(row[idx["FullName"]]))
        person = {"name": normalize_name(row[idx["FullName"]])}
        timeline = build_timeline(row, idx)
        if timeline is not None:
            person["timeline"] = timeline
            report["timeline_counts"]["with_any_timeline"] += 1
            if timeline.get("birthYear") is not None:
                report["timeline_counts"]["with_birth_year"] += 1
            if timeline.get("deathYear") is not None:
                report["timeline_counts"]["with_death_year"] += 1
            if timeline.get("startYear") is not None:
                report["timeline_counts"]["with_start_year"] += 1
            if timeline.get("endYear") is not None:
                report["timeline_counts"]["with_end_year"] += 1
            if timeline.get("startYear") is not None and timeline.get("endYear") is not None:
                report["timeline_counts"]["with_lifespan_range"] += 1
            if timeline.get("estimatedBirthYear"):
                report["timeline_counts"]["estimated_birth_year"] += 1
        for loc in all_locations_by_person[key]:
            report["location_counts"].setdefault(loc.kind, 0)
            if (
                loc.lat is None
                or loc.lon is None
                or ("placeholder_coords" in loc.issues and not loc.changed)
                or ("street_only_far_from_dc" in loc.issues and not loc.changed)
            ):
                report["dropped_locations"] += 1
                manual_review.append(
                    {
                        "name": loc.person_name,
                        "guid": loc.guid,
                        "row_number": loc.row_number,
                        "kind": loc.kind,
                        "issue": ";".join(loc.issues) or "dropped_unresolved",
                        "address": loc.display_address() or "",
                        "lat": loc.current_lat,
                        "lon": loc.current_lon,
                        "query_attempts": " | ".join(loc.attempted_queries),
                    }
                )
                continue

            if loc.changed:
                report["changed_locations"] += 1

            for issue in loc.issues:
                report["issue_counts"][issue] += 1
            for note in loc.notes:
                report["notes_counts"][note] += 1

            person[loc.kind] = {
                "address": loc.display_address(),
                "lat": round(loc.lat, 10),
                "lon": round(loc.lon, 10),
            }
            report["location_counts"][loc.kind] += 1

            unresolved = [
                issue
                for issue in loc.issues
                if issue in {
                    "city_at_state_centroid",
                    "georgetown_historical_outlier",
                    "street_at_city_centroid",
                    "street_only_far_from_dc",
                    "outside_expected_state",
                    "conflicting_address_coords",
                }
            ]
            if unresolved:
                manual_review.append(
                    {
                        "name": loc.person_name,
                        "guid": loc.guid,
                        "row_number": loc.row_number,
                        "kind": loc.kind,
                        "issue": ";".join(unresolved),
                        "address": loc.display_address() or "",
                        "lat": loc.lat,
                        "lon": loc.lon,
                        "query_attempts": " | ".join(loc.attempted_queries),
                    }
                )
        output_people.append(person)

    report["issue_counts"] = dict(report["issue_counts"])
    report["notes_counts"] = dict(report["notes_counts"])
    report["timeline_counts"] = dict(report["timeline_counts"])

    timeline_stats = {
        "min_year": None,
        "max_year": None,
        "people_with_any_timeline": report["timeline_counts"].get("with_any_timeline", 0),
        "people_with_start_year": report["timeline_counts"].get("with_start_year", 0),
        "people_with_lifespan_range": report["timeline_counts"].get("with_lifespan_range", 0),
        "estimated_birth_year": report["timeline_counts"].get("estimated_birth_year", 0),
    }
    for person in output_people:
        timeline = person.get("timeline")
        if not timeline:
            continue
        for year in [timeline.get("startYear"), timeline.get("endYear"), timeline.get("birthYear"), timeline.get("deathYear")]:
            if year is None:
                continue
            if timeline_stats["min_year"] is None or year < timeline_stats["min_year"]:
                timeline_stats["min_year"] = year
            if timeline_stats["max_year"] is None or year > timeline_stats["max_year"]:
                timeline_stats["max_year"] = year

    dataset = {
        "people": output_people,
        "source": workbook_path.name,
        "cleaning": {
            "based_on": input_json_path.name,
            "script": str(Path(__file__).name),
            "live_geocode": geocode_live,
        },
        "timeline": timeline_stats,
    }
    return dataset, report, manual_review


def write_manual_review(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["name", "guid", "row_number", "kind", "issue", "address", "lat", "lon", "query_attempts"],
        )
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def main() -> None:
    parser = argparse.ArgumentParser(description="Clean the MZ/FUBS atlas geocoded dataset.")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--input-json", type=Path, default=DEFAULT_INPUT_JSON)
    parser.add_argument("--output-json", type=Path, default=DEFAULT_OUTPUT_JSON)
    parser.add_argument("--state-polygons", type=Path, default=DEFAULT_STATE_POLYGONS)
    parser.add_argument("--cache", type=Path, default=DEFAULT_CACHE)
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT)
    parser.add_argument("--manual-review", type=Path, default=DEFAULT_MANUAL_REVIEW)
    parser.add_argument("--geocode-live", action="store_true")
    parser.add_argument("--sleep-seconds", type=float, default=1.0)
    args = parser.parse_args()

    dataset, report, manual_review = build_clean_dataset(
        workbook_path=args.workbook,
        input_json_path=args.input_json,
        state_polygons_path=args.state_polygons,
        geocode_live=args.geocode_live,
        cache_path=args.cache,
        sleep_seconds=args.sleep_seconds,
    )

    args.output_json.write_text(json.dumps(dataset, indent=2))
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(json.dumps(report, indent=2, sort_keys=True))
    write_manual_review(args.manual_review, manual_review)

    print(json.dumps({
        "people": len(dataset["people"]),
        "location_counts": report["location_counts"],
        "changed_locations": report["changed_locations"],
        "dropped_locations": report["dropped_locations"],
        "query_stats": report["query_stats"],
        "timeline": dataset.get("timeline", {}),
        "manual_review_rows": len(manual_review),
    }, indent=2))


if __name__ == "__main__":
    main()
